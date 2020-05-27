"""
    目标地址：http://www.chinatax.gov.cn/chinatax/n810346/n810825/index.html
    1. 采集
    采集以下栏目：
        增值税、消费税的前两页数据
    每个栏目的：
        标题、发文日期、文号

    2. 保存
        将所有信息保存到`税务局.xlsx`文件，依据栏目名创建数据表，每个栏目的数据保存到对应的数据表
"""
import openpyxl
import requests
import json
import time
import pprint

url = 'http://www.chinatax.gov.cn/api/query'
params = {
    'siteCode': 'bm29000fgk',
    'tab': 'all',
    'key': '9A9C42392D397C5CA6C1BF07E2E0AA6F'
}
headers = {
    'Host': 'www.chinatax.gov.cn',
    'Origin': 'http://www.chinatax.gov.cn',
    'Referer': 'http://www.chinatax.gov.cn/chinatax/n810346/n810825/index.html',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/81.0.4044.138 Safari/537.36',
    'Cookie': '_Jo0OQK=6DD6068EE5E472AE4BF5734392231D83B421EFCFF3CBCC1153659DA3A421E6579D391F85E1593E4154259074AB3BCA9A475271A43FADEF418C1BFC4F163C587ABA201C83797AB67C7627EF4EAE0C6A2C5647EF4EAE0C6A2C56491A0E8A83004FA3FGJ1Z1Qg==; yfx_c_g_u_id_10003701=_ck20052000082415776872774379191; yfx_f_l_v_t_10003701=f_t_1589904504571__r_t_1589952760120__v_t_1589963407466__r_c_1; CPS_SESSION=2B84D1C175AA8067AC0F3EF04B5B3776'
}


def get_data(con, page):
    form_data = {
        'timeOption': '0',
        'C6': f'{con}',
        'page': f'{page}',
        'pageSize': '10',
        'keyPlace': '1',
        'sort': 'dateDesc',
        'qt': '*'
    }
    return form_data


def get_mes(con, page):
    response = requests.post(url=url,
                             params=params,
                             data=get_data(con, page),
                             headers=headers)
    res_dict = json.loads(response.content.decode(encoding='utf-8-sig'))
    # pprint.pprint(res_dict)
    res = res_dict['resultList']
    wb = openpyxl.load_workbook('税务局.xlsx', read_only=False)
    if con not in wb.sheetnames:
        i = 1
        sheet = wb.create_sheet(con)
        for data in res:
            dreTitle = data['dreTitle']
            publishTime = data['publishTime']
            documentNumber = data['customHs']['DOCNOVAL']
            # print(dreTitle, publishTime, documentNumber)
            sheet.cell(row=i, column=1).value = dreTitle
            sheet.cell(row=i, column=2).value = publishTime
            sheet.cell(row=i, column=3).value = documentNumber
            i += 1
            wb.save('税务局.xlsx')
    else:
        # pass
        i = (11 * (page - 1))
        table = wb.get_sheet_by_name(con)
        # print(table.title)
        for data in res:
            dreTitle = data['dreTitle']
            publishTime = data['publishTime']
            documentNumber = data['customHs']['DOCNOVAL']
            # print(dreTitle, publishTime, documentNumber)
            table.cell(row=i, column=1).value = dreTitle
            table.cell(row=i, column=2).value = publishTime
            table.cell(row=i, column=3).value = documentNumber
            i += 1
            wb.save('税务局.xlsx')


if __name__ == '__main__':
    wb = openpyxl.Workbook()
    wb.save('税务局.xlsx')
    for page in range(1, 3):
        get_mes('增值税', page)
        time.sleep(5)
        print("succeed")
    for page in range(1, 3):
        get_mes('消费税', page)
        time.sleep(5)
        print("succeed")
